from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root=Tk()
root.title("Sistema de registro automotriz")
root.geometry("1250x700")
root.config(bg=background)

file=pathlib.Path('registro.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] = "Registro No."
    sheet['B1'] = "Rut"
    sheet['C1'] = "Nombre cliente"
    sheet['D1'] = "Apellido paterno"
    sheet['E1'] = "Apellido materno"
    sheet['F1'] = "Genero"
    sheet['G1'] = "Marca vehiculo"
    sheet['H1'] = "Patente"
    sheet['I1'] = "Modelo"
    sheet['J1'] = "Año"
    sheet['K1'] = "Fecha registro"
    sheet['L1'] = "Estado"

    file.save('registro.xlsx')

#exit
def Exit():
    root.destroy()


#Mostrar imagen
def showimage():
    global filename
    global img
    filename= filedialog.askopenfilename(initialdir=os.getcwd(),
                                         title="Select image file",filetypes=(("JPG file", "*.jpg"),
                                                                              ("PNG file","*.png"),
                                                                              ("All files","*.txt")))
    img = (Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image_names=photo2

#Registrar
def registrar():
    file=openpyxl.load_workbook('registro.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row, column=1).value

    try:
        Registracion.set(max_row_value+1)
    except:
        Registracion.set("1")


#Clear
def Clear():
    global img
    Rut.set('')
    Name.set('')
    apPaterno.set('')
    apMaterno.set('')
    Marca.set('')
    Patente.set('')
    Modelo.set('')
    Year.set('')
    Fecha.set('')
    Estado.set("Selecciona opcion")

    registrar()

    saveButton.config( state= 'normal')

    img1=PhotoImage(file='imagenes/Captura3.PNG')
    lbl.config(img=img1)
    lbl.image_names=img1

    img=""

def Save():
    R1=Registracion.get()
    N1=Name.get()
    E1=Estado.get()
    try:
        G1=genero
    except:
        messagebox.showerror("error", "Seleciona el genero")

    R2=Rut.get()
    A1=apPaterno.get()
    A2=apMaterno.get()
    M1=Marca.get()
    P1=Patente.get()
    M2=Modelo.get()
    Y1=Year.get()
    F1=Fecha.get()


    if N1=="" or E1=="Selecciona opcion" or R2=="" or A1=="" or A2=="" or M1=="" or P1=="" or M2=="" or Y1=="" or F1=="":
        messagebox.showerror("error", "Llenar todos los campos")
    else:
        file=openpyxl.load_workbook('registro.xlsx')
        sheet=file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=R2)
        sheet.cell(column=3, row=sheet.max_row, value=N1)
        sheet.cell(column=4, row=sheet.max_row, value=A1)
        sheet.cell(column=5, row=sheet.max_row, value=A2)
        sheet.cell(column=6, row=sheet.max_row, value=G1)
        sheet.cell(column=7, row=sheet.max_row, value=M1)
        sheet.cell(column=8, row=sheet.max_row, value=P1)
        sheet.cell(column=9, row=sheet.max_row, value=M2)
        sheet.cell(column=10, row=sheet.max_row, value=Y1)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=E1)

        file.save(r'registro.xlsx')
        
        try:
            img.save("imagenes2/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","La imagen no esta disponible")

        messagebox.showinfo("info","Datos ingresados exitosamente")

        Clear()

        registrar()

#Buscar
def search():
    text = Search.get()

    Clear()
    saveButton.config(state="no habilitado")
    file=openpyxl.load_workbook("registro.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalido", "Numero invalido")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    Registracion.set(x1)
    Rut.set(x2)
    Name.set(x3)
    apPaterno.set(x4)
    apMaterno.set(x5)

    if x6=='F':
        R2.select()
    else:
        R1.select()
    Marca.set(x7)
    Patente.set(x8)
    Modelo.set(x9)
    Year.set(x10)
    Fecha.set(x11)
    Estado.set(x12)

def Update():
    pass



#genero
def seleccion():
    global genero
    value=radio.get()
    if value==1:
        genero="F"
    else:
        genero="M"


Label (root,text="Email: taller_automotriz@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP,fill=X)
Label (root,text="REGISTRO AUTOMOTRIZ", width=10, height=2, bg="#c36464", fg="#fff", font='arial 20 bold').pack(side=TOP,fill=X)

Search=StringVar()
Entry(root,textvariable=Search,width=15, bd=2, font="arial 20").place(x=820, y=70)
imageicon3=PhotoImage(file="imagenes/Captura1.png")
Srch=Button(root, text="Buscar", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold", command=search)
Srch.place(x=1060, y=66)

imageicon4=PhotoImage(file="imagenes/Captura2.PNG")
Update_button=Button(root,image=imageicon4, bg="#c36464", command=Update)
Update_button.place(x=110,y=64)

Label(root,text="Registro No", font="arial 13", fg=framebg, bg=background).place(x=30,y=150)
Label(root,text="Fecha", font="arial 13", fg=framebg, bg=background).place(x=500,y=150)

Registracion=IntVar()
Date= StringVar()

reg_entry = Entry(root, textvariable=Registracion, width=15,font="arial 10")
reg_entry.place(x=160,y=150)


registrar()

today = date.today()
d1=today.strftime("%d/%m/%Y")
date_entry =Entry(root,textvariable=Date,width=15, font="arial 10")
date_entry.place(x=600,y=150)

Date.set(d1)

obj=LabelFrame(root,text="Datos cliente", font=20, bd=2,width= 900, bg=framebg, fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Rut", font="arial 13", bg=framebg, fg=framefg).place(x=30,y=50)
Label(obj,text="Nombre cliente", font="arial 13", bg=framebg, fg=framefg).place(x=30,y=100)
Label(obj,text="Apellido paterno", font="arial 13", bg=framebg, fg=framefg).place(x=30,y=150)


Label(obj,text="Apellido materno", font="arial 13", bg=framebg, fg=framefg).place(x=500,y=50)
Label(obj,text="Genero", font="arial 13", bg=framebg, fg=framefg).place(x=500,y=100)

Rut=StringVar()
rut_entry=Entry(obj,textvariable=Rut, width=20, font="arial 10")
rut_entry.place(x=160,y=50)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160,y=100)


apPaterno=StringVar()
AP_entry=Entry(obj,textvariable=apPaterno, width=20, font="arial 10")
AP_entry.place(x=160,y=150)

apMaterno=StringVar()
AM_entry=Entry(obj,textvariable=apMaterno, width=20, font="arial 10")
AM_entry.place(x=630,y=50)

radio=IntVar()
R1=Radiobutton(obj,text="F", variable=radio, value=1, bg=framebg, fg=framefg,command=seleccion)
R1.place(x=570,y=100)


R2=Radiobutton(obj,text="M", variable=radio, value=2, bg=framebg, fg=framefg,command=seleccion)
R2.place(x=630,y=100)





obj2=LabelFrame(root,text="Datos vehiculo", font=20, bd=2,width= 900, bg=framebg, fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)


Label(obj2,text="Marca vehiculo",font="arial 13", bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Patente",font="arial 13", bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj2,text="Modelo",font="arial 13", bg=framebg,fg=framefg).place(x=30,y=150)

Marca=StringVar()
marca_entry=Entry(obj2,textvariable=Marca, width=20, font="arial 10")
marca_entry.place(x=160,y=50)

Patente=StringVar()
pat_entry=Entry(obj2,textvariable=Patente, width=20, font="arial 10")
pat_entry.place(x=160,y=100)

Modelo=StringVar()
model_entry=Entry(obj2,textvariable=Modelo, width=20, font="arial 10")
model_entry.place(x=160,y=150)




Label(obj2,text="Año",font="arial 13", bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Fecha registro",font="arial 13", bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj2,text="Estado",font="arial 13", bg=framebg,fg=framefg).place(x=500,y=150)

Year=StringVar()
year_entry=Entry(obj2,textvariable=Year, width=20, font="arial 10")
year_entry.place(x=630,y=50)

Fecha=StringVar()
fech_entry=Entry(obj2,textvariable=Fecha, width=20, font="arial 10")
fech_entry.place(x=630,y=100)


Estado=Combobox(obj2, values=["Entrando a taller", "En mantencion", "Operativo"], font="Roboto 10",width=17,state="r")
Estado.place(x=630,y=150)
Estado.set("Selecciona opcion")


f=Frame(root,bd=3, bg="black", width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="imagenes/Captura3.PNG")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#botones

Button(root,text="Cargar",width=19,height=2,font="arial 12",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Guardar",width=19,height=2,font="arial 12",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="Resetear",width=19,height=2,font="arial 12",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Salir",width=19,height=2,font="arial 12",bg="grey", command=Exit).place(x=1000,y=610)












































root.mainloop()



